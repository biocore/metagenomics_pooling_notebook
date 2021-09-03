import unittest
import tempfile
import sys
from datetime import datetime

from metapool.igm import IGMManifest
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


class IGMManifestTests(unittest.TestCase):
    def test_raises_error_with_missing_properties(self):
        m = IGMManifest()
        with self.assertRaisesRegex(ValueError, 'project_number cannot be '
                                    'empty, you need to set a value'):
            m.write()

    def test_raises_error_with_missing_properties_single(self):
        m = IGMManifest()
        m.project_number = 1000
        m.task_number = 3
        m.number_of_lanes = 1
        m.pools = ['TMI_AGP_300_302_304_308']
        with self.assertRaisesRegex(ValueError, 'number_of_samples cannot be '
                                    'empty, you need to set a value'):
            m.write()

    def test_default_values(self):
        m = IGMManifest()

        self.assertEqual(m.submission_date,
                         datetime.strftime(datetime.today(), '%m/%d/%y'))
        self.assertEqual(m.institute, 'Knight Lab')
        self.assertEqual(m.pi_name, 'Dr. Knight')
        self.assertEqual(m.pi_email, 'mackenzie.m.bryant@gmail.com')
        self.assertEqual(m.contact_name, 'MacKenzie Bryant')
        self.assertEqual(m.contact_email, 'mackenzie.m.bryant@gmail.com')

        self.assertIsNone(m.project_number)
        self.assertIsNone(m.task_number)
        self.assertEqual(m.platform, 'NovaSeq S4')
        self.assertEqual(m.run_type, 'PE150')
        self.assertEqual(m.custom_primer,
                         'No-Standard Illumina Primers are fine')

        self.assertIsNone(m.number_of_samples)
        self.assertEqual(m.number_of_lanes, '1')

        self.assertIsNone(m._pools)

        self.assertTrue(isinstance(m._workbook, Workbook))

    def test_default_path(self):
        m = IGMManifest()
        m.project_number = 123
        m.task_number = 2
        m.number_of_samples = 96
        m.number_of_lanes = 1
        m.pools = ['spooky_scary_skeletons']
        m.pi_name = 'R. Caballero'

        exp = (datetime.strftime(datetime.today(), '%Y_%m_%d') +
               '_Caballero_spooky_scary_skeletons_Manifest_2021.xlsx')
        self.assertEqual(m._default_path(), exp)

    def test_write_and_load(self):
        m = IGMManifest()
        m.submission_date = '01/01/00'
        m.project_number = 1000
        m.task_number = 3
        m.number_of_samples = 384
        m.number_of_lanes = 1
        m.pools = ['Scooby_dooby_doo']

        with tempfile.NamedTemporaryFile('w+', suffix='.xlsx') as tmp:
            m.write(tmp.name)
            exp_name = tmp.name

            sheet = load_workbook(tmp.name)['Sample Information']

            self.assertEqual(sheet['B2'].value, '01/01/00')
            self.assertEqual(sheet['B3'].value, 'Knight Lab')
            self.assertEqual(sheet['B4'].value, 'Dr. Knight')
            self.assertEqual(sheet['B5'].value, 'mackenzie.m.bryant@gmail.com')
            self.assertEqual(sheet['B6'].value, 'MacKenzie Bryant')
            self.assertEqual(sheet['B7'].value, 'mackenzie.m.bryant@gmail.com')
            self.assertEqual(sheet['B12'].value, 1000)
            self.assertEqual(sheet['B13'].value, 3)
            self.assertEqual(sheet['B18'].value, 'NovaSeq S4')
            self.assertEqual(sheet['B19'].value, 'PE150')
            self.assertEqual(sheet['B20'].value,
                             'No-Standard Illumina Primers are fine')
            self.assertEqual(sheet['B22'].value, 384)
            self.assertEqual(sheet['B23'].value, 1)

            self.assertEqual(sheet['D1'].value, '150x8x8x150')
            self.assertEqual(sheet['D1'].fill,
                             PatternFill(fill_type='solid', fgColor='FFFF00'))

            self.assertEqual(sheet['A25'].value, 'Scooby_dooby_doo')
            self.assertEqual(sheet['B25'].value, 'Scooby_dooby_doo')
            self.assertEqual(sheet['C25'].value, 500)
            self.assertEqual(sheet['D25'].value, 'KHP')

        observed = sys.stdout.getvalue().strip()
        self.assertEqual(observed, 'Saving manifest to %s' % exp_name)

    def test_write_pools(self):
        m = IGMManifest()
        m.submission_date = '01/01/00'
        m.project_number = 1000
        m.pools = ['the_plate', 'iron_plate', 'maiden_plate']

        self.assertEqual(m.submission_date, '01/01/00')
        self.assertEqual(m.project_number, 1000)

        self.assertEqual(m._sheet['A25'].value, 'the_plate')
        self.assertEqual(m._sheet['B25'].value, 'the_plate')
        self.assertEqual(m._sheet['C25'].value, 500)
        self.assertEqual(m._sheet['D25'].value, 'KHP')

        self.assertEqual(m._sheet['A26'].value, 'iron_plate')
        self.assertEqual(m._sheet['B26'].value, 'iron_plate')
        self.assertEqual(m._sheet['C26'].value, 500)
        self.assertEqual(m._sheet['D26'].value, 'KHP')

        self.assertEqual(m._sheet['A27'].value, 'maiden_plate')
        self.assertEqual(m._sheet['B27'].value, 'maiden_plate')
        self.assertEqual(m._sheet['C27'].value, 500)
        self.assertEqual(m._sheet['D27'].value, 'KHP')

        m.pools = ['scooby_dooby_doo']

        self.assertEqual(m._sheet['A25'].value, 'scooby_dooby_doo')
        self.assertEqual(m._sheet['B25'].value, 'scooby_dooby_doo')
        self.assertEqual(m._sheet['C25'].value, 500)
        self.assertEqual(m._sheet['D25'].value, 'KHP')

        self.assertIsNone(m._sheet['A26'].value)
        self.assertIsNone(m._sheet['B26'].value)
        self.assertIsNone(m._sheet['C26'].value)
        self.assertIsNone(m._sheet['D26'].value)

        self.assertIsNone(m._sheet['A27'].value)
        self.assertIsNone(m._sheet['B27'].value)
        self.assertIsNone(m._sheet['C27'].value)
        self.assertIsNone(m._sheet['D27'].value)


if __name__ == '__main__':
    assert not hasattr(sys.stdout, "getvalue")
    unittest.main(module=__name__, buffer=True, exit=False)
