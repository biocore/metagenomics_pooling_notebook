import os

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

_IGM_ROW_OFFSET = 25

_CONTACT = 'MacKenzie Bryant'
_EMAIL = 'mackenzie.m.bryant@gmail.com'
_INSTITUTE = 'Knight Lab'
_PI_NAME = 'Dr. Knight'
_PLATFORM = 'NovaSeq S4'
_RUN_TYPE = 'PE150'
_CUSTOM_PRIMER = 'No-Standard Illumina Primers are fine'


def _property_maker(obj, name, cell, default):
    """Create a setter/getter for an attribute and associate it with a cell

    Parameters
    ----------
    obj: IGMManifest
        The instance that needs the property.
    name: str
        The name of the property for the instance.
    cell: str
        The sheet's cell where the property should be written to.
    default: str, int, float or None
        Default value for the property.
    """
    def fget(self):
        return getattr(self, '_' + name)

    def fset(self, value):
        if value is not None:
            self._sheet[cell] = value
        setattr(self, '_' + name, value)

    # set the default value
    fset(obj, default)

    setattr(type(obj), name, property(fget, fset))


class IGMManifest(object):
    """Interface to write an IGM manifest file

    Attributes
    ----------
    submission_date: str
        When the sample sheet is submitted. Defaults to today's date.
    institute: str
        Institution or laboratory.  Defaults to "Knight Lab".
    pi_name: str
        Name of the PI. Defaults to "Dr. Knight".
    pi_email: str
        Email for the PI. Defaults to mackenzie.m.bryant@gmail.com
    contact_name: str
        Contact's name. Defaults to MacKenzie Bryant.
    project_number: int
        Project number. Defaults to 2004033.
    task_number: int
        Tasks number. Defaults to 1.
    platform: str
        Sequencing platform. Defaults to "NovaSeq S4".
    run_type: str
        Sequencing run type. Defaults to "PE150".
    custom_primer: str
        Notes about the primers in the run. Defaults to "No-Standard Illumina
        Primers are fine".
    number_of_samples: int
        Samples in the run.
    number_of_lanes: int
        Samples in the lane.
    pools: list of str
        List of pool names to include in the sheet.
    """
    def __init__(self):
        self._workbook = self._load_igm_template()
        self._sheet = self._workbook['Sample Information']

        _property_maker(self, 'submission_date', 'B2',
                        datetime.strftime(datetime.today(), '%m/%d/%y'))
        _property_maker(self, 'institute', 'B3', _INSTITUTE)
        _property_maker(self, 'pi_name', 'B4', _PI_NAME)
        _property_maker(self, 'pi_email', 'B5', _EMAIL)
        _property_maker(self, 'contact_name', 'B6', _CONTACT)
        _property_maker(self, 'contact_email', 'B7', _EMAIL)

        _property_maker(self, 'project_number', 'B12', 2004033)
        _property_maker(self, 'task_number', 'B13', 1)

        _property_maker(self, 'platform', 'B18', _PLATFORM)
        _property_maker(self, 'run_type', 'B19', _RUN_TYPE)
        _property_maker(self, 'custom_primer', 'B20', _CUSTOM_PRIMER)

        _property_maker(self, 'number_of_samples', 'B22', None)
        _property_maker(self, 'number_of_lanes', 'B23', '1')

        self._pools = None

        # default value requested by IGM
        self._sheet['D1'] = '150x8x8x150'
        self._sheet['D1'].fill = PatternFill(fill_type='solid',
                                             fgColor='FFFF00')

    @property
    def pools(self):
        return self._pools

    @pools.setter
    def pools(self, value):
        """List of pools to write in the manifest"""
        if self._pools is not None:
            for i, pool in enumerate(self._pools):
                row = str(_IGM_ROW_OFFSET + i)

                del self._sheet['A' + row]
                del self._sheet['B' + row]
                del self._sheet['C' + row]
                del self._sheet['D' + row]
        if value is not None:
            for i, pool in enumerate(value):
                row = str(_IGM_ROW_OFFSET + i)

                self._sheet['A' + row] = pool
                self._sheet['B' + row] = pool
                self._sheet['C' + row] = 500
                self._sheet['D' + row] = 'KHP'

        self._pools = value

    def __str__(self):
        out = (
            f'Date of Sample Submission: {self.submission_date}\n'
            f'Institute/Company Name: {self.institute}\n'
            f'PI Name: {self.pi_name}\n'
            f'PI Email: {self.pi_email}\n'
            f'Contact Name: {self.contact_name}\n'
            f'Contact Email: {self.contact_email}\n'
            f'Project Number: {self.project_number}\n'
            f'Task Number: {self.task_number}\n'
            f'Platform: {self.platform}\n'
            f'Run Type: {self.run_type}\n'
            'Custom Primer? (Provide more info in comments box): '
            f'{self.custom_primer}\n'
            f'Total number of Samples: {self.number_of_samples}\n'
            'Total number of Lanes OR Total Reads Required: '
            f'{self.number_of_lanes}\n\n'
            'Sample Name\tPool Name\tLibrary Size (bp)\tLibrary Prep Method\n'
        )

        if self.pools:
            pools = '\n'.join(f'{p}\t{p}\t500\tKHP'for p in self.pools)
        else:
            pools = ''

        return out + pools

    def write(self, path=None):
        """Write the manifest to disk

        Parameters
        ----------
        path: str, optional
            Path where to write the manifest. If `None`, IGM's default format
            is used (YYYY_MM_DD_PI_Sequencing_Runs_Manifest_2021.xlsx).

        Raises
        ------
        ValueError
            If a required attribute is missing.
        """
        required = ['project_number', 'task_number', 'platform',
                    'number_of_samples', 'number_of_lanes', 'pools']
        for prop in required:
            if getattr(self, prop) is None:
                raise ValueError('%s cannot be empty, you need to set a value'
                                 % prop)

        path = self._default_path() if path is None else path
        print('Saving manifest to %s' % path)
        print(str(self))

        self._workbook.save(path)

    def _default_path(self):
        path = (datetime.strftime(datetime.today(), '%Y_%m_%d_') +
                # should be the last name
                self.pi_name.split(' ')[-1] + '_' +
                # all the pools with spaces replaced for underscores
                '_'.join([p.replace(' ', '_') for p in self._pools]) +
                '_Manifest_2021.xlsx')
        return path

    def _load_igm_template(self):
        """Helper function to load IGM's manifest template"""
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data',
                            'template.xlsx')
        return load_workbook(path)
